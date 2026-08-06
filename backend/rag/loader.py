"""
文档加载器 — 支持多种文件格式（无需 unstructured 依赖）
"""
import os
from typing import List
from langchain_core.documents import Document


class DocumentLoader:
    """多格式文档加载器"""

    SUPPORTED_TYPES = ["pdf", "docx", "txt", "md", "csv", "xlsx"]

    @staticmethod
    def load(file_path: str) -> List[Document]:
        """根据文件类型加载文档，返回 LangChain Document 列表"""
        ext = os.path.splitext(file_path)[1].lower().lstrip(".")

        if ext not in DocumentLoader.SUPPORTED_TYPES:
            raise ValueError(f"不支持的文件类型: .{ext}，支持的类型: {DocumentLoader.SUPPORTED_TYPES}")

        if ext == "pdf":
            from langchain_community.document_loaders import PyPDFLoader
            loader = PyPDFLoader(file_path)

        elif ext == "docx":
            from langchain_community.document_loaders import Docx2txtLoader
            loader = Docx2txtLoader(file_path)

        elif ext in ("txt", "md"):
            from langchain_community.document_loaders import TextLoader
            loader = TextLoader(file_path, encoding="utf-8")

        elif ext == "csv":
            from langchain_community.document_loaders import CSVLoader
            loader = CSVLoader(file_path, encoding="utf-8")

        elif ext == "xlsx":
            # 使用 openpyxl 直接读取
            docs = DocumentLoader._load_xlsx(file_path)
            filename = os.path.basename(file_path)
            for doc in docs:
                doc.metadata["filename"] = filename
                doc.metadata["file_type"] = "xlsx"
            return docs

        else:
            raise ValueError(f"不支持的文件类型: .{ext}")

        docs = loader.load()

        # 添加来源元数据
        filename = os.path.basename(file_path)
        for doc in docs:
            doc.metadata["filename"] = filename
            doc.metadata["file_type"] = ext

        return docs

    @staticmethod
    def _load_xlsx(file_path: str) -> List[Document]:
        """使用 openpyxl 直接加载 Excel 文件"""
        import openpyxl
        docs = []
        wb = openpyxl.load_workbook(file_path, read_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            # 将每行转为文本
            lines = []
            for row in rows:
                line = "\t".join([str(cell) if cell is not None else "" for cell in row])
                lines.append(line)
            content = "\n".join(lines)
            docs.append(Document(
                page_content=content,
                metadata={"source": file_path, "sheet": sheet_name},
            ))
        wb.close()
        return docs
